from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill

from scraper.models import Business


COLUMN_LABELS = {
    "google_rank": "Rang Google",
    "name": "Nom",
    "category": "Catégorie",
    "address": "Adresse",
    "postal_code": "Code postal",
    "locality": "Localité",
    "city": "Ville recherchée",
    "phone": "Téléphone",
    "email": "Email pro",
    "website": "Site web",
    "managers": "Dirigeant(s)",
    "vat_number": "Numéro TVA",
    "bce_number": "Numéro BCE",
    "legal_form": "Forme juridique",
    "bce_status": "Statut BCE",
    "creation_date": "Date de création",
    "capital": "Capital",
    "establishments_count": "Nb établissements",
    "nace_activities": "Activités (NACE)",
    "bce_match_score": "Score match TVA",
    "bce_match_warning": "Note enrichissement TVA",
    "nbb_url": "Comptes annuels (BNB)",
    "nbb_year": "Dernier exercice BNB",
    "nbb_revenue": "Chiffre d'affaires",
    "nbb_equity": "Fonds propres",
    "nbb_employees": "Effectif",
    "companyweb_url": "Fiche CompanyWeb",
    "companyweb_score": "Score solvabilité",
    "rating": "Note Google",
    "reviews_count": "Nombre d'avis",
    "hours": "Horaires",
    "gmaps_url": "Lien Google Maps",
    "plus_code": "Plus Code",
    "query": "Requête",
    "already_seen": "Déjà vue",
    "first_seen": "Vue pour la 1re fois",
}

COLUMN_ORDER = [
    "google_rank", "name", "category",
    "address", "postal_code", "locality", "city",
    "phone", "email", "website",
    "managers",
    "vat_number", "bce_number", "legal_form", "bce_status",
    "creation_date", "capital", "establishments_count", "nace_activities",
    "nbb_url", "nbb_year", "nbb_revenue", "nbb_equity", "nbb_employees",
    "companyweb_url", "companyweb_score",
    "bce_match_score", "bce_match_warning",
    "rating", "reviews_count", "hours",
    "gmaps_url", "plus_code", "query",
    "already_seen", "first_seen",
]


def to_dataframe(businesses: list[Business]) -> pd.DataFrame:
    rows = [b.to_dict() for b in businesses]
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    cols = [c for c in COLUMN_ORDER if c in df.columns]
    df = df[cols].rename(columns=COLUMN_LABELS)
    return df


def to_excel_bytes(businesses: list[Business], sheet_name: str = "Prospects") -> bytes:
    df = to_dataframe(businesses)
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]

        # En-tête : fond bleu marine, texte blanc
        header_fill = PatternFill("solid", fgColor="1E293B")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Largeur auto
        for col_cells in worksheet.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            col_letter = col_cells[0].column_letter
            worksheet.column_dimensions[col_letter].width = min(max(length + 2, 12), 55)

        worksheet.freeze_panes = "A2"

        # Surligner les 1er et 2e du classement Google
        rank_col = None
        for idx, cell in enumerate(worksheet[1], start=1):
            if cell.value == COLUMN_LABELS["google_rank"]:
                rank_col = idx
                break

        if rank_col is not None:
            gold = PatternFill("solid", fgColor="FEF3C7")
            silver = PatternFill("solid", fgColor="E5E7EB")
            for row in worksheet.iter_rows(min_row=2):
                rank_value = row[rank_col - 1].value
                if rank_value == 1:
                    for c in row:
                        c.fill = gold
                elif rank_value == 2:
                    for c in row:
                        c.fill = silver

    return buffer.getvalue()
